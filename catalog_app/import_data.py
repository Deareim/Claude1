#!/usr/bin/env python3
"""Import Excel Integration Catalog into SQLite database."""

import openpyxl
import sqlite3
import re
import os
import sys

EXCEL_FILE = os.path.join(os.path.dirname(__file__), '..', 'IntegrationCatalog_Consolidated_20260320_1701.xlsx')
DB_FILE = os.path.join(os.path.dirname(__file__), 'catalog.db')

SHEET_TABLE_MAP = {
    'SAP CPI':                      ('cpi',            'SAP CPI Interfaces'),
    'H (For HTTP Based Services)':  ('http_services',  'HTTP Services'),
    'F (For FILE Based Services)':  ('file_services',  'File Services'),
    'E (For EVENT Based Services)': ('event_services', 'Event Services'),
    'S (For SOAP Based Services)':  ('soap_services',  'SOAP Services'),
    'R (For RFC Based Services)':   ('rfc_services',   'RFC Services'),
    'D (For IDOC Based Services)':  ('idoc_services',  'IDoc Services'),
    'W (For WSS Based Services)':   ('wss_services',   'WSS Services'),
    'M (For MAIL Based Service)':   ('mail_services',  'Mail Services'),
}


def clean_col_name(name: str) -> str:
    if not name:
        return 'col_unknown'
    name = str(name).strip()
    name = re.sub(r'[\s\n\r\t]+', '_', name)
    name = re.sub(r'[^\w]', '_', name)
    name = re.sub(r'_+', '_', name)
    name = name.strip('_').lower()
    if len(name) > 55:
        name = name[:55].rstrip('_')
    return name or 'col_unknown'


def make_unique_headers(raw_headers: list) -> list:
    seen = set()
    counter = {}
    result = []
    for h in raw_headers:
        clean = clean_col_name(h)
        if clean not in seen:
            seen.add(clean)
            result.append(clean)
        else:
            counter[clean] = counter.get(clean, 1) + 1
            candidate = f"{clean}_{counter[clean]}"
            while candidate in seen:
                counter[clean] += 1
                candidate = f"{clean}_{counter[clean]}"
            seen.add(candidate)
            result.append(candidate)
    return result


def import_data():
    print(f"Source : {os.path.abspath(EXCEL_FILE)}")
    print(f"Target : {DB_FILE}\n")

    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    if os.path.exists(DB_FILE):
        os.remove(DB_FILE)
        print("Removed existing database.\n")

    conn = sqlite3.connect(DB_FILE)

    conn.execute('''CREATE TABLE _tables (
        table_name   TEXT PRIMARY KEY,
        display_name TEXT,
        sheet_name   TEXT
    )''')

    conn.execute('''CREATE TABLE _column_labels (
        table_name  TEXT,
        col_name    TEXT,
        label       TEXT,
        PRIMARY KEY (table_name, col_name)
    )''')

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    for sheet_name, (table_name, display_name) in SHEET_TABLE_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP : '{sheet_name}' not found in workbook")
            continue

        ws = wb[sheet_name]

        raw_headers = [cell.value for cell in ws[1]]
        # Trim trailing None headers
        while raw_headers and raw_headers[-1] is None:
            raw_headers.pop()

        if not raw_headers:
            print(f"  SKIP : No headers in '{sheet_name}'")
            continue

        col_names = make_unique_headers(raw_headers)
        n_cols = len(col_names)

        col_defs = ', '.join([f'"{c}" TEXT' for c in col_names])
        conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" (id INTEGER PRIMARY KEY AUTOINCREMENT, {col_defs})')

        conn.execute('INSERT INTO _tables VALUES (?, ?, ?)', [table_name, display_name, sheet_name])

        for col, raw in zip(col_names, raw_headers):
            if raw:
                label = str(raw).split('\n')[0].strip()[:120]
                conn.execute('INSERT OR REPLACE INTO _column_labels VALUES (?, ?, ?)', [table_name, col, label])

        placeholders  = ', '.join(['?'] * n_cols)
        quoted_cols   = ', '.join([f'"{c}"' for c in col_names])
        rows_inserted = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            values = []
            for i in range(n_cols):
                v = row[i] if i < len(row) else None
                values.append(str(v) if v is not None else None)
            conn.execute(f'INSERT INTO "{table_name}" ({quoted_cols}) VALUES ({placeholders})', values)
            rows_inserted += 1

        conn.commit()
        print(f"  OK  : {sheet_name:<45}  →  {table_name:<18}  ({rows_inserted} rows, {n_cols} cols)")

    wb.close()
    conn.close()
    print(f"\nDatabase ready: {DB_FILE}")


if __name__ == '__main__':
    import_data()
